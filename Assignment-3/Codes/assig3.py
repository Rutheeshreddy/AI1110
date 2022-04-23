
from openpyxl  import Workbook, load_workbook

wb = load_workbook('assig3.xlsx')
ws = wb.active

N=0
for cols in range(1,5):
        for rows in range(2,7):
                char=chr(65+cols)
                N=N+ws[char+str(rows)].value
print('Total number of families is {}'.format(N))

p1 = ws['D4'].value/N 
print('Probability of families earning Rs.10000 – 13000 per \
 month and owning 2 vehicles is {0:.3f}'.format(p1))

p2 = ws['C6'].value/N
print('Probability of families earning Rs.16000 or more per \
 month and owning 1 vehicle is {0:.3f}'.format(p2))

p3 = ws['B2'].value/N
print('Probability of families earning Rs.7000 or less per \
 month and owning no vehicle is {0:.4f}'.format(p3))

p4 = ws['E5'].value/N
print('Probability of families earning Rs.13000 – 16000 per \
 month and owning more than 2 vehicles is {0:.4f}'.format(p4))

n=0
for cols in range(1,3):
        for rows in range(2,7):
                char=chr(65+cols)
                n=n+ws[char+str(rows)].value
print(n)
p5=n/N
print('Probability of families owning not more than 1 vehicle is {0:.3f}'.format(p5))

wb.close()